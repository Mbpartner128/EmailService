# report.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import config

C = {
    "header":    "1F3864",
    "sent":      "C6EFCE",
    "warn":      "FFEB9C",
    "error":     "FFC7CE",
    "skip":      "D9D9D9",
    "personal":  "DDEBF7",
    "sheet_hdr": "2F5496",
}

def _fill(h): return PatternFill("solid", start_color=h)
def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")
def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _row_color(r: dict) -> str:
    if r.get("duplicate"):     return C["skip"]
    if r.get("skip_status"):   return C["skip"]
    if not r.get("valid"):     return C["error"]
    if r.get("mx_fail"):       return C["error"]
    if r.get("is_personal") and config.BLOCK_PERSONAL_EMAIL:
        return C["skip"]
    if r.get("rate_limited"):  return C["warn"]
    if r.get("sent"):          return C["sent"]
    if r.get("is_personal"):   return C["personal"]
    return "FFFFFF"


def write_report(results: list[dict], path: str = None) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    Path(config.OUTPUT_DIR).mkdir(exist_ok=True)
    out = path or f"{config.OUTPUT_DIR}/result_{ts}.xlsx"

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Results ──────────────────────────────────
    ws = wb.active
    ws.title = "All Results"
    ws.freeze_panes = "A2"

    headers = ["State", "Name", "Email", "Status (original)",
               "Format", "MX Record", "Personal", "Send Result", "Notes"]
    ws.append(headers)
    for i, cell in enumerate(ws[1], 1):
        cell.font      = _font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _fill(C["header"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()
    ws.row_dimensions[1].height = 22

    for r in results:
        row = [
            r.get("_sheet", ""),
            r.get("name", ""),
            r.get("email", ""),
            r.get("status", ""),
            "✅" if r.get("valid") else "❌",
            ("✅" if r.get("mx_ok") else "❌") if r.get("mx_ok") is not None else "–",
            "⚠️" if r.get("is_personal") else "–",
            "✅ Sent" if r.get("sent") else (
                "🧪 Dry run" if r.get("dry_run") else (
                    "🚫 Failed" if r.get("send_failed") else "–"
                )
            ),
            r.get("notes", ""),
        ]
        ws.append(row)
        bg = _row_color(r)
        last = ws.max_row
        for cell in ws[last]:
            cell.fill   = _fill(bg)
            cell.font   = _font(size=9)
            cell.border = _thin_border()

    # Column widths
    col_widths = [14, 30, 38, 16, 8, 10, 10, 12, 38]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Per-State Summary ────────────────────────────
    ws2 = wb.create_sheet("By State")
    ws2.freeze_panes = "A2"

    sum_headers = ["State", "Total", "Valid Email", "No Email",
                   "Invalid Format", "No MX", "Personal", "Duplicates",
                   "Status Skipped", "Sent", "Send Failed"]
    ws2.append(sum_headers)
    for cell in ws2[1]:
        cell.font      = _font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _fill(C["sheet_hdr"])
        cell.alignment = Alignment(horizontal="center")
        cell.border    = _thin_border()

    by_state = defaultdict(list)
    for r in results:
        by_state[r.get("_sheet", "Unknown")].append(r)

    for state, rows in by_state.items():
        total       = len(rows)
        valid       = sum(1 for r in rows if r.get("valid"))
        no_email    = sum(1 for r in rows if not r.get("email"))
        invalid_fmt = sum(1 for r in rows if r.get("email") and not r.get("valid"))
        no_mx       = sum(1 for r in rows if r.get("mx_fail"))
        personal    = sum(1 for r in rows if r.get("is_personal"))
        duplicates  = sum(1 for r in rows if r.get("duplicate"))
        skipped     = sum(1 for r in rows if r.get("skip_status"))
        sent        = sum(1 for r in rows if r.get("sent"))
        failed      = sum(1 for r in rows if r.get("send_failed"))
        ws2.append([state, total, valid, no_email, invalid_fmt,
                    no_mx, personal, duplicates, skipped, sent, failed])
        last = ws2.max_row
        for cell in ws2[last]:
            cell.border = _thin_border()
            cell.font   = _font(size=9)

    # Totals row
    n = len(by_state) + 1
    ws2.append([
        "TOTAL",
        f"=SUM(B2:B{n})", f"=SUM(C2:C{n})", f"=SUM(D2:D{n})",
        f"=SUM(E2:E{n})", f"=SUM(F2:F{n})", f"=SUM(G2:G{n})",
        f"=SUM(H2:H{n})", f"=SUM(I2:I{n})", f"=SUM(J2:J{n})",
        f"=SUM(K2:K{n})",
    ])
    for cell in ws2[ws2.max_row]:
        cell.font   = _font(bold=True, size=10)
        cell.fill   = _fill("E2EFDA")
        cell.border = _thin_border()

    for i, w in enumerate([20,8,10,10,14,8,10,10,16,8,12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Ready to Send ────────────────────────────────
    ws3 = wb.create_sheet("Ready to Send")
    ws3.freeze_panes = "A2"
    ws3.append(["State", "Name", "Email", "Personal Domain?"])
    for cell in ws3[1]:
        cell.font  = _font(bold=True, color="FFFFFF")
        cell.fill  = _fill(C["header"])
        cell.border = _thin_border()

    sendable = [r for r in results if r.get("sendable")]
    for r in sendable:
        ws3.append([
            r.get("_sheet", ""),
            r.get("name", ""),
            r.get("email", ""),
            "Yes" if r.get("is_personal") else "No",
        ])
        last = ws3.max_row
        bg = C["personal"] if r.get("is_personal") else "FFFFFF"
        for cell in ws3[last]:
            cell.fill   = _fill(bg)
            cell.font   = _font(size=9)
            cell.border = _thin_border()

    for i, w in enumerate([14, 30, 38, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 4: Legend ───────────────────────────────────────
    ws4 = wb.create_sheet("Legend")
    legend = [
        ("Color", "Meaning"),
        ("🟢 Green",  "Email sent successfully"),
        ("🔵 Blue",   "Valid personal/free domain when BLOCK_PERSONAL_EMAIL = False"),
        ("🟡 Yellow", "Not sent because hourly or daily send limit was reached"),
        ("🔴 Red",    "Invalid format or no MX record — skipped"),
        ("⬜ Gray",   "Skipped due to Status column, duplicate email, or blocked personal/free domain"),
        ("⬜ White",  "Valid corporate email, not yet sent"),
        ("", ""),
        ("Column", "Meaning"),
        ("Format ✅/❌", "Email address passes basic format check"),
        ("MX Record ✅/❌", "Domain DNS has a mail server (can receive email)"),
        ("Personal ⚠️", f"Domain is in personal list: {', '.join(sorted(config.PERSONAL_DOMAINS)[:6])}…"),
        ("Personal blocking", f"BLOCK_PERSONAL_EMAIL = {config.BLOCK_PERSONAL_EMAIL}"),
        ("Sent ✅", "Email was successfully delivered via SMTP"),
        ("Dry run 🧪", "Email was selected but not actually sent because DRY_RUN = True"),
        ("Status Skipped", f"Original status column matched: {config.SKIP_STATUSES}"),
    ]
    for r in legend:
        ws4.append(list(r))
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 65
    ws4[1][0].font = ws4[1][1].font = _font(bold=True)
    ws4[9][0].font = ws4[9][1].font = _font(bold=True)

    wb.save(out)
    print(f"\n📄 Report saved → {out}")
    return out
